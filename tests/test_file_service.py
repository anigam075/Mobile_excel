import tempfile
import unittest
from pathlib import Path

from app.models.workbook import Sheet, Workbook, column_name
from app.services.file_service import load_workbook, save_workbook


class WorkbookTests(unittest.TestCase):
    def test_column_name(self):
        self.assertEqual(column_name(0), "A")
        self.assertEqual(column_name(25), "Z")
        self.assertEqual(column_name(26), "AA")

    def test_insert_and_delete_selected_cells(self):
        workbook = Workbook(
            sheets=[Sheet(name="Data", rows=[["A", "B"], ["1", "2"], ["3", "4"]])],
            file_type="csv",
        )

        workbook.insert_row_after(0)
        self.assertEqual(workbook.active_sheet.get_cell(1, 0), "")
        self.assertEqual(workbook.active_sheet.get_cell(2, 0), "1")

        workbook.add_cell_below(1, 0)
        self.assertEqual(workbook.active_sheet.get_cell(1, 0), "")
        self.assertEqual(workbook.active_sheet.get_cell(2, 0), "")
        self.assertEqual(workbook.active_sheet.get_cell(3, 0), "1")

        workbook.delete_cell(2, 0)
        self.assertEqual(workbook.active_sheet.get_cell(2, 0), "1")
        self.assertEqual(workbook.active_sheet.get_cell(3, 0), "3")

        workbook.delete_row(2)
        self.assertEqual(workbook.active_sheet.get_cell(2, 0), "3")

    def test_csv_round_trip(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "sample.csv"
            workbook = Workbook(
                sheets=[Sheet(name="Sheet1", rows=[["Metric", "Amount"], ["", ""]])],
                file_type="csv",
            )
            workbook.set_cell(1, 0, "Revenue")
            workbook.set_cell(1, 1, "1200")

            save_workbook(workbook, path)
            loaded = load_workbook(path)

            self.assertEqual(loaded.active_sheet.get_cell(1, 0), "Revenue")
            self.assertEqual(loaded.active_sheet.get_cell(1, 1), "1200")

    def test_xlsx_round_trip(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "sample.xlsx"
            workbook = Workbook(
                sheets=[
                    Sheet(
                        name="Data",
                        rows=[["Metric", "Amount"], ["", ""]],
                        freeze_top_row=True,
                    )
                ],
                file_type="xlsx",
            )
            workbook.set_cell(1, 0, "Cost")
            workbook.set_cell(1, 1, "900")

            save_workbook(workbook, path)
            loaded = load_workbook(path)

            self.assertEqual(loaded.sheets[0].name, "Data")
            self.assertTrue(loaded.sheets[0].freeze_top_row)
            self.assertEqual(loaded.active_sheet.get_cell(1, 0), "Cost")
            self.assertEqual(loaded.active_sheet.get_cell(1, 1), "900")

            stored = load_workbook(path, storage_dir=directory)
            try:
                self.assertTrue(stored.sheets[0].freeze_top_row)
            finally:
                stored.close(remove=True)


if __name__ == "__main__":
    unittest.main()
