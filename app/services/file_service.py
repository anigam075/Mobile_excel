import csv
import posixpath
import xml.etree.ElementTree as ElementTree
import zipfile
from pathlib import Path
from uuid import uuid4

from app.models.workbook import Sheet, Workbook
from app.services.workbook_store import StoredWorkbook, WorkbookStore


class FileServiceError(Exception):
    pass


class OperationCancelled(FileServiceError):
    pass


SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
IMPORT_BATCH_SIZE = 250


def load_workbook(path, storage_dir=None, progress_callback=None, cancel_event=None):
    file_path = Path(path)
    if not file_path.exists():
        raise FileServiceError(f"File does not exist: {file_path}")

    suffix = file_path.suffix.lower()
    if storage_dir:
        if suffix == ".csv":
            return _load_csv_stored(file_path, storage_dir, progress_callback, cancel_event)
        if suffix == ".xlsx":
            return _load_xlsx_stored(file_path, storage_dir, progress_callback, cancel_event)
    elif suffix == ".csv":
        return _load_csv(file_path)
    elif suffix == ".xlsx":
        return _load_xlsx(file_path)
    raise FileServiceError("Only .csv and .xlsx files are supported.")


def save_workbook(workbook, path, progress_callback=None, cancel_event=None):
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise FileServiceError("Save path must end with .csv or .xlsx.")

    file_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = file_path.with_name(
        f".{file_path.stem}.mobilexl-{uuid4().hex}.tmp{suffix}"
    )
    try:
        if suffix == ".csv":
            _save_csv(workbook, temporary_path, progress_callback, cancel_event)
        else:
            _save_xlsx(workbook, temporary_path, progress_callback, cancel_event)
        temporary_path.replace(file_path)
    except FileServiceError:
        raise
    except OSError as exc:
        raise FileServiceError(f"Could not replace saved file: {exc}") from exc
    finally:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass


def _load_csv(file_path):
    try:
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [[cell for cell in row] for row in csv.reader(handle)]
    except UnicodeDecodeError:
        with file_path.open("r", encoding="latin-1", newline="") as handle:
            rows = [[cell for cell in row] for row in csv.reader(handle)]
    except OSError as exc:
        raise FileServiceError(f"Could not read CSV file: {exc}") from exc

    if not rows:
        rows = [[""]]
    return Workbook(
        sheets=[Sheet(name=file_path.stem or "Sheet1", rows=rows)],
        file_type="csv",
        path=str(file_path),
    )


def _load_csv_stored(file_path, storage_dir, progress_callback, cancel_event):
    last_unicode_error = None
    for encoding in ("utf-8-sig", "latin-1"):
        store = WorkbookStore.create(storage_dir)
        try:
            sheet_id = store.add_sheet(file_path.stem or "Sheet1", 0)
            imported_rows = 0
            batch = []
            with file_path.open("r", encoding=encoding, newline="") as handle:
                for row in csv.reader(handle):
                    _check_cancelled(cancel_event)
                    batch.append(row)
                    if len(batch) >= IMPORT_BATCH_SIZE:
                        store.append_rows(sheet_id, batch, imported_rows)
                        imported_rows += len(batch)
                        batch.clear()
                        _notify_progress(
                            progress_callback,
                            imported_rows,
                            None,
                            f"Imported {imported_rows:,} rows",
                        )
                if batch:
                    store.append_rows(sheet_id, batch, imported_rows)
                    imported_rows += len(batch)

            _check_cancelled(cancel_event)
            _notify_progress(
                progress_callback,
                imported_rows,
                imported_rows,
                f"Imported {imported_rows:,} rows",
            )
            return StoredWorkbook(
                store,
                store.sheet_metadata(),
                file_type="csv",
                path=str(file_path),
            )
        except UnicodeDecodeError as exc:
            last_unicode_error = exc
            store.close(remove=True)
        except OperationCancelled:
            store.close(remove=True)
            raise
        except OSError as exc:
            store.close(remove=True)
            raise FileServiceError(f"Could not read CSV file: {exc}") from exc
        except Exception as exc:
            store.close(remove=True)
            raise FileServiceError(f"Could not import CSV file: {exc}") from exc

    raise FileServiceError(f"Could not decode CSV file: {last_unicode_error}")


def _save_csv(workbook, file_path, progress_callback=None, cancel_event=None):
    try:
        file_path.parent.mkdir(parents=True, exist_ok=True)
        total_rows = workbook.active_sheet.row_count
        with file_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            for row_index, row in enumerate(_iter_sheet_rows(workbook.active_sheet), start=1):
                _check_cancelled(cancel_event)
                writer.writerow(row)
                if row_index % IMPORT_BATCH_SIZE == 0:
                    _notify_progress(
                        progress_callback,
                        row_index,
                        total_rows,
                        f"Saved {row_index:,} of {total_rows:,} rows",
                    )
    except OperationCancelled:
        raise
    except OSError as exc:
        raise FileServiceError(f"Could not save CSV file: {exc}") from exc


def _load_xlsx(file_path):
    try:
        from openpyxl import load_workbook as openpyxl_load_workbook
    except ImportError as exc:
        raise FileServiceError("openpyxl is required to open .xlsx files.") from exc

    try:
        source = openpyxl_load_workbook(file_path, read_only=True)
    except Exception as exc:
        raise FileServiceError(f"Could not read Excel file: {exc}") from exc

    try:
        frozen_top_rows = _xlsx_frozen_top_rows(file_path)
        sheets = []
        for worksheet in source.worksheets:
            rows = []
            for row in worksheet.iter_rows(values_only=True):
                rows.append(["" if value is None else str(value) for value in row])
            if not rows:
                rows = [[""]]
            sheets.append(
                Sheet(
                    name=worksheet.title,
                    rows=rows,
                    freeze_top_row=frozen_top_rows.get(worksheet.title, False),
                )
            )
    finally:
        source.close()

    if not sheets:
        sheets = [Sheet(name="Sheet1", rows=[[""]])]

    return Workbook(sheets=sheets, file_type="xlsx", path=str(file_path))


def _load_xlsx_stored(file_path, storage_dir, progress_callback, cancel_event):
    try:
        from openpyxl import load_workbook as openpyxl_load_workbook
    except ImportError as exc:
        raise FileServiceError("openpyxl is required to open .xlsx files.") from exc

    try:
        source = openpyxl_load_workbook(file_path, read_only=True)
    except Exception as exc:
        raise FileServiceError(f"Could not read Excel file: {exc}") from exc

    store = WorkbookStore.create(storage_dir)
    frozen_top_rows = _xlsx_frozen_top_rows(file_path)
    imported_rows = 0
    total_rows = sum(worksheet.max_row or 0 for worksheet in source.worksheets)
    try:
        for position, worksheet in enumerate(source.worksheets):
            _check_cancelled(cancel_event)
            sheet_id = store.add_sheet(
                worksheet.title,
                position,
                freeze_top_row=frozen_top_rows.get(worksheet.title, False),
            )
            sheet_row = 0
            batch = []
            for row in worksheet.iter_rows(values_only=True):
                _check_cancelled(cancel_event)
                batch.append(["" if value is None else str(value) for value in row])
                if len(batch) >= IMPORT_BATCH_SIZE:
                    store.append_rows(sheet_id, batch, sheet_row)
                    sheet_row += len(batch)
                    imported_rows += len(batch)
                    batch.clear()
                    _notify_progress(
                        progress_callback,
                        imported_rows,
                        total_rows,
                        f"Imported {imported_rows:,} of {total_rows:,} rows",
                    )
            if batch:
                store.append_rows(sheet_id, batch, sheet_row)
                imported_rows += len(batch)

        if not source.worksheets:
            store.add_sheet("Sheet1", 0)
        _check_cancelled(cancel_event)
        return StoredWorkbook(
            store,
            store.sheet_metadata(),
            file_type="xlsx",
            path=str(file_path),
        )
    except OperationCancelled:
        store.close(remove=True)
        raise
    except Exception as exc:
        store.close(remove=True)
        raise FileServiceError(f"Could not import Excel file: {exc}") from exc
    finally:
        source.close()


def _save_xlsx(workbook, file_path, progress_callback=None, cancel_event=None):
    try:
        from openpyxl import Workbook as OpenpyxlWorkbook
    except ImportError as exc:
        raise FileServiceError("openpyxl is required to save .xlsx files.") from exc

    target = OpenpyxlWorkbook(write_only=True)
    total_rows = sum(sheet.row_count for sheet in workbook.sheets)
    saved_rows = 0
    for sheet in workbook.sheets:
        worksheet = target.create_sheet(title=sheet.name[:31] or "Sheet")
        if sheet.freeze_top_row:
            worksheet.freeze_panes = "A2"
        for row in _iter_sheet_rows(sheet):
            _check_cancelled(cancel_event)
            worksheet.append(row)
            saved_rows += 1
            if saved_rows % IMPORT_BATCH_SIZE == 0:
                _notify_progress(
                    progress_callback,
                    saved_rows,
                    total_rows,
                    f"Saved {saved_rows:,} of {total_rows:,} rows",
                )

    try:
        file_path.parent.mkdir(parents=True, exist_ok=True)
        target.save(file_path)
    except OperationCancelled:
        raise
    except OSError as exc:
        raise FileServiceError(f"Could not save Excel file: {exc}") from exc


def _iter_sheet_rows(sheet):
    if hasattr(sheet, "iter_rows"):
        return sheet.iter_rows()
    return iter(sheet.normalized_rows())


def _xlsx_frozen_top_rows(file_path):
    relationship_attribute = (
        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    )
    try:
        with zipfile.ZipFile(file_path) as archive:
            workbook_root = ElementTree.parse(archive.open("xl/workbook.xml")).getroot()
            relationships_root = ElementTree.parse(
                archive.open("xl/_rels/workbook.xml.rels")
            ).getroot()
            targets = {
                relationship.attrib["Id"]: relationship.attrib["Target"]
                for relationship in relationships_root
                if "Id" in relationship.attrib and "Target" in relationship.attrib
            }

            frozen = {}
            for sheet in workbook_root.iter():
                if sheet.tag.rsplit("}", 1)[-1] != "sheet":
                    continue
                relationship_id = sheet.attrib.get(relationship_attribute)
                target = targets.get(relationship_id)
                if not target:
                    continue
                worksheet_path = _xlsx_relationship_path(target)
                frozen[sheet.attrib.get("name", "")] = _worksheet_xml_freezes_top_row(
                    archive,
                    worksheet_path,
                )
            return frozen
    except (KeyError, OSError, ElementTree.ParseError, zipfile.BadZipFile):
        return {}


def _xlsx_relationship_path(target):
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _worksheet_xml_freezes_top_row(archive, worksheet_path):
    with archive.open(worksheet_path) as worksheet_xml:
        for _event, element in ElementTree.iterparse(worksheet_xml, events=("start",)):
            element_name = element.tag.rsplit("}", 1)[-1]
            if element_name == "pane":
                try:
                    return float(element.attrib.get("ySplit", 0)) >= 1
                except ValueError:
                    return False
            if element_name == "sheetData":
                break
    return False


def _check_cancelled(cancel_event):
    if cancel_event is not None and cancel_event.is_set():
        raise OperationCancelled("Operation cancelled.")


def _notify_progress(callback, current, total, message):
    if callback is not None:
        callback(current, total, message)
